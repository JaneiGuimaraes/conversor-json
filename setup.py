import pandas as pd
from tkinter import Tk, filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def selecionar_arquivo(titulo):
    """Abre uma janela para selecionar um arquivo."""
    root = Tk()
    root.withdraw()
    caminho = filedialog.askopenfilename(
        title=titulo,
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    return caminho

def ajustar_formato_excel(caminho_arquivo):
    """Aplica formatação automática às células do arquivo Excel."""
    wb = load_workbook(caminho_arquivo)
    ws = wb.active

    # Ajusta largura das colunas e formata células
    for col in ws.columns:
        max_length = 0
        coluna = col[0].column_letter  # Pega a letra da coluna (A, B, C...)

        # Encontra o tamanho máximo do conteúdo na coluna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        # Define a largura da coluna (com margem adicional)
        largura_ajustada = (max_length + 5) if max_length > 10 else 15
        ws.column_dimensions[coluna].width = largura_ajustada

        # Aplica quebra de texto e alinhamento central
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    wb.save(caminho_arquivo)

def filtrar_e_salvar(arquivo1, arquivo2, arquivo_saida):
    """Filtra os nomes presentes no Arquivo 1 e pega os dados do Arquivo 2."""
    try:
        # Lê os arquivos
        df1 = pd.read_excel(arquivo1)
        df2 = pd.read_excel(arquivo2)
        
        # Verifica se a coluna 'Nome' existe
        if 'Nome' not in df1.columns or 'Nome' not in df2.columns:
            raise ValueError("A coluna 'Nome' não foi encontrada em um dos arquivos.")
        
        # Pega os nomes do Arquivo 1 (sem espaços extras)
        nomes_arquivo1 = set(df1['Nome'].str.strip().dropna())
        
        # Filtra o Arquivo 2 mantendo apenas linhas cujos nomes estão no Arquivo 1
        df_resultado = df2[df2['Nome'].str.strip().isin(nomes_arquivo1)]
        
        # Salva o resultado
        df_resultado.to_excel(arquivo_saida, index=False)
        
        # Aplica formatação ao arquivo gerado
        ajustar_formato_excel(arquivo_saida)
        
        return len(df1), len(df2), len(df_resultado)
    
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro: {str(e)}")
        return None

if __name__ == "__main__":
    print("=== Filtrador de Arquivos Excel ===")
    
    # Seleciona os arquivos
    arquivo1 = selecionar_arquivo("Selecione o PRIMEIRO arquivo (base de nomes)")
    if not arquivo1:
        print("Operação cancelada.")
        exit()
    
    arquivo2 = selecionar_arquivo("Selecione o SEGUNDO arquivo (dados para filtragem)")
    if not arquivo2:
        print("Operação cancelada.")
        exit()
    
    # Define o nome do arquivo de saída
    arquivo_saida = arquivo2.replace(".xlsx", "_FILTRADO.xlsx")
    
    # Confirmação
    confirmar = messagebox.askyesno(
        "Confirmar",
        f"Gerar arquivo com base nos nomes de '{arquivo1}' usando dados de '{arquivo2}'?\n"
        f"Arquivo de saída: {arquivo_saida}"
    )
    
    if not confirmar:
        print("Processo cancelado.")
        exit()
    
    # Processa os dados
    resultado = filtrar_e_salvar(arquivo1, arquivo2, arquivo_saida)
    
    if resultado:
        total_arquivo1, total_arquivo2, total_resultado = resultado
        mensagem = (
            f"✅ Concluído!\n\n"
            f"🔹 Linhas no Arquivo 1 (nomes de referência): {total_arquivo1}\n"
            f"🔹 Linhas no Arquivo 2 (dados brutos): {total_arquivo2}\n"
            f"🔹 Linhas no Arquivo Final (filtrado): {total_resultado}\n\n"
            f"📂 Salvo em: {arquivo_saida}"
        )
        messagebox.showinfo("Sucesso", mensagem)
        print(mensagem)