import json
import openpyxl
from openpyxl.styles import Alignment, Font
from tkinter import Tk, filedialog, messagebox
import sys
import os

def selecionar_arquivo():
    root = Tk()
    root.withdraw()
    return filedialog.askopenfilename(
        title="Selecione o arquivo JSON",
        filetypes=[("JSON", "*.json")]
    )

def extrair_referencia(produto):
    """Função dedicada para extrair a referência interna de qualquer estrutura"""
    # Caso 1: Campo direto no produto
    if 'internalReference' in produto:
        return produto['internalReference']
    
    # Caso 2: Dentro de 'productInformation'
    if 'productInformation' in produto and isinstance(produto['productInformation'], dict):
        return produto['productInformation'].get('internalReference', 'N/A')
    
    # Caso 3: Dentro de estruturas aninhadas
    for key, value in produto.items():
        if isinstance(value, dict) and 'internalReference' in value:
            return value['internalReference']
    
    return 'N/A'  # Padrão se não encontrar

def formatar_descricao(descricoes):
    """Processa descrições com formatação técnica completa"""
    if not descricoes:
        return "Descrição não disponível"
    
    partes = []
    for desc in descricoes if isinstance(descricoes, list) else [descricoes]:
        if isinstance(desc, dict):
            texto = desc.get('value', '')
            if texto:
                linhas = [f"• {linha[2:].strip()}" if linha.startswith('- ') else linha for linha in texto.split('\n')]
                partes.append('\n'.join(linhas))
    return '\n\n'.join(partes)

def formatar_opcionais(opcionais):
    """Extrai todos os opcionais com hierarquia visual"""
    if not opcionais:
        return "Nenhum opcional cadastrado"
    
    textos = []
    if isinstance(opcionais, dict):
        if 'budgetPage' in opcionais:
            textos.append(f"• Incluso no orçamento: {'Sim' if opcionais['budgetPage'] else 'Não'}")
        if 'productPage' in opcionais:
            textos.append(f"• Visível no catálogo: {'Sim' if opcionais['productPage'] else 'Não'}")
    
    if isinstance(opcionais, list):
        for grupo in opcionais:
            if isinstance(grupo, dict) and 'name' in grupo:
                opcoes = [f"  ▸ {op['name']}" for op in grupo.get('optionals', []) if isinstance(op, dict)]
                if opcoes:
                    textos.append(f"• {grupo['name']}:\n" + '\n'.join(opcoes))
    
    return '\n'.join(textos) or "Nenhum opcional cadastrado"

def gerar_planilha(json_path):
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            dados = json.load(f)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Produtos"
        
        # Cabeçalhos
        cabecalhos = ["PRODUTO", "DESCRIÇÃO COMPLETA", "REF. INTERNA", "OPCIONAIS"]
        ws.append(cabecalhos)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # Processar produtos
        for produto in dados if isinstance(dados, list) else [dados]:
            ws.append([
                produto.get('name', 'Sem nome').strip(),
                formatar_descricao(produto.get('descriptions', [])),
                extrair_referencia(produto),  # Agora extrai corretamente
                formatar_opcionais(produto.get('optionals', []))
            ])
        
        # Ajuste de layout
        colunas = {
            'A': 35,  # Nome
            'B': 70,  # Descrição
            'C': 20,  # Referência
            'D': 40   # Opcionais
        }
        for col, width in colunas.items():
            ws.column_dimensions[col].width = width
        
        estilo = Alignment(wrap_text=True, vertical='top')
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = estilo
            ws.row_dimensions[row[0].row].height = max(60, len(str(row[1].value).split('\n')) * 15)
        
        # Salvar
        output_path = os.path.splitext(json_path)[0] + "_FINAL.xlsx"
        wb.save(output_path)
        
        messagebox.showinfo(
            "Sucesso!",
            f"Planilha gerada com:\n"
            f"- Referências internas extraídas\n"
            f"- Todos dados técnicos\n"
            f"- Formatação profissional\n\n"
            f"Salvo em: {output_path}"
        )
    
    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao processar:\n{e}")
        sys.exit(1)

if __name__ == "__main__":
    if arquivo := selecionar_arquivo():
        gerar_planilha(arquivo)